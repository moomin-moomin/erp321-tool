import asyncio
from os import listdir
from typing import cast

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from xlrd import Book, open_workbook
from xlrd.sheet import Sheet


def _is_valid_xlsx_file(file_name: str):
    # 以“~$”开头的文件是Excel软件自动生成的备份文件，跳过处理
    if file_name.startswith("~$"):
        return False
    if "【推单】" not in file_name:
        return False
    return file_name.endswith(".xlsx") or file_name.endswith(".xls")


def get_input_xlsl_file_paths():
    """
    获取同文件夹下文件名中包含“【推单】”关键字的文件名
    """
    children = listdir()
    return [child for child in children if _is_valid_xlsx_file(child)]


def workbook_to_dicts(workbook: Workbook):
    """
    将workbook对象转换为一个字典列表。
    与dicts_to_workbook是一组互逆转换。

    例如：
    这个表格
    | a | b | c |
    | 1 | 2 | 3 |
    | 4 | 5 | 6 |
    将会被转换为
    [
        {"a": 1, "b": 2, "c": 3},
        {"a": 4, "b": 5, "c": 6},
    ]
    """
    sheetname = workbook.sheetnames[0]
    sheet = workbook[sheetname]
    dicts = [
        {str(sheet.cell(row=1, column=cell.column).value): cell.value for cell in row}
        for index, row in enumerate(sheet.rows)
        if index
    ]
    return dicts


def dicts_to_workbook(rows: list[dict]):
    """
    将字典列表转换为一个workbook对象，用于后续保存为xlsx文件。
    与workbook_to_dicts是一组互逆转换。

    例如：
    这个字典列表
    [
        {"a": 1, "b": 2, "c": 3},
        {"a": 4, "b": 5, "c": 6},
    ]
    将会被转换为
    | a | b | c |
    | 1 | 2 | 3 |
    | 4 | 5 | 6 |
    """
    workbook = Workbook()
    sheetname = workbook.sheetnames[0]
    sheet = workbook[sheetname]

    example_row = rows[0]
    column_names = example_row.keys()

    current_column = 1
    for column_name in column_names:
        cell = sheet.cell(row=1, column=current_column)
        cell.value = column_name
        current_column += 1

    current_row = 2
    for row in rows:
        for index, column_name in enumerate(column_names):
            cell = sheet.cell(row=current_row, column=index + 1)
            cell.value = row[column_name]
        current_row += 1
    return workbook


def sheet_to_dict(sheet: Worksheet):
    """
    将xlsx文件中的一个sheet转换为二维dict。其中第一维的key为表格的列名称，第二维的key为表格的行名称。
    如果单元格中的内容包含逗号“，”或空格“ ”，对应dict的值会被分隔为list。
    空单元格会被丢弃

    例如：
    这个表格
    |   | a | b | c |
    | x | 1 | 2 | 3 |
    | y | 4 | 5 |6 7|
    将会被转换为
    {
        "a": {"x": "1", y: "4"},
        "b": {"x": "2", y: "5"},
        "c": {"x": "3", y: ["6", "7"]},
    }
    """
    sheet_dict: dict[str, dict[str, str | list[str]]] = {}
    for column in sheet.columns:
        column = cast(tuple[Cell], column)
        column = [cell for cell in column if cell.value is not None]
        for cell in column:
            value = str(cell.value)
            if cell.row == 1 and cell.column > 1:
                sheet_dict[value] = sheet_dict.get(value, {})
            if cell.row > 1 and cell.column > 1:
                column_name = str(sheet.cell(row=cell.row, column=1).value)
                shop_name = str(sheet.cell(row=1, column=cell.column).value)
                if "," in value:
                    value = value.split(",")
                elif " " in value:
                    value = value.split(" ")
                sheet_dict[shop_name][column_name] = value
    return sheet_dict


def syncify(function, callback=None):
    """
    将异步函数转为同步并带回调的函数。
    如果异步函数执行出现异常，异常会作为参数传给回调函数。
    """

    def syncified(*args, **kwargs):
        """
        由传入的异步函数“function”转换后的同步执行函数
        """
        try:
            asyncio.run(function(*args, **kwargs))
            if callable(callback):
                callback(None)
        except Exception as exception:
            if callable(callback):
                callback(exception)
            else:
                raise exception

    return syncified


def _copy_xls_sheet_to_xlsx_sheet(xls_sheet: Sheet, xlsx_sheet: Worksheet):
    for row in range(xls_sheet.nrows):
        for column, xls_cell_value in enumerate(xls_sheet.row_values(row)):
            xlsx_cell = xlsx_sheet.cell(row=row + 1, column=column + 1)
            xlsx_cell.value = xls_cell_value


def _xls_workbook_to_xlsx_workbook(xls_workbook: Book):
    xlsx_workbook = Workbook()
    for sheetname in xlsx_workbook.sheetnames:
        xlsx_workbook.remove(xlsx_workbook[sheetname])
    for xls_sheet in xls_workbook.sheets():
        xlsx_sheet = xlsx_workbook.create_sheet()
        _copy_xls_sheet_to_xlsx_sheet(xls_sheet, xlsx_sheet)
    return xlsx_workbook


def load_xlsx_or_xls_workbook(path: str):
    if path.endswith(".xlsx"):
        return load_workbook(path)

    if path.endswith(".xls"):
        workbook = open_workbook(path)
        return _xls_workbook_to_xlsx_workbook(workbook)

    raise RuntimeError("不支持的文件格式")
